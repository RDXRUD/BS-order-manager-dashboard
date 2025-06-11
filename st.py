import streamlit as st
import os
import sys
from openpyxl import load_workbook
import win32com.client
import pythoncom
import base64
from openpyxl.styles import Border, Side
from datetime import date
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill,Alignment
from openpyxl.utils import get_column_letter

def convert_excel_to_pdf(excel_file, pdf_file):
    pythoncom.CoInitialize()
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False

    workbook = excel.Workbooks.Open(excel_file)
    sheet = workbook.Sheets(1)
    
    try:
        # Adjust page setup to fit content to PDF
        sheet.PageSetup.Zoom = False
        sheet.PageSetup.FitToPagesWide = 1  # Fit all columns to one page width
        sheet.PageSetup.FitToPagesTall = False  # Allow multiple pages for rows if needed
        
        sheet.Select()
        workbook.ExportAsFixedFormat(0, pdf_file)
            
    except Exception as e:
        print(f"Error: {e}")
    finally:
        workbook.Close(False)
        excel.Quit()
        pythoncom.CoUninitialize()

def fetch_products(file,date,company,tc,tl,orderNO,emails):
    # print("hi")
    # os.startfile(file)
    name,extension=company.split(".")
    workbook = load_workbook(file)
    
    sheet_names = workbook.sheetnames 
    # st.write(sheet_names)
    # st.write(f"Sheet names: {sheet_names}")

    if len(sheet_names) > 1:
        for sheet_name in sheet_names[1:]:
            del workbook[sheet_name]

    sheet_names = workbook.sheetnames 
    # st.write(f"Remaining sheet names: {sheet_names}")

    sheet = workbook[sheet_names[0]]

    findHead=False
    find_qty = False
    qty_index = None
    serial_count = 0
    blank_line = 0
    find_date=False
    date_column=None
    headingRow=None
    name_index=None
    countStr=None
    findStar=False
    
    small_index=None
    medium_index=None
    large_index=None
    xl_index=None
    uni_index=None
    
    head=[]

    # Find 'Qty.' column
    for row in sheet.iter_rows():
        for cell in row:
            # st.write(cell.value)
            if cell.value is not None and isinstance(cell.value, str) and ("qty" in cell.value.strip().lower() or "name" in cell.value.strip().lower() or "size" in cell.value.strip().lower()):
                # st.write(cell.value)
                findHead=True
                
                # find_qty = True
                # qty_column = cell.column
                headingRow=cell.row 
                # st.write(headingRow)
                # st.write(f"'qty' found at Row: {cell.row}, Column: {cell.column}")
                break
        if findHead:
            break
    
    # if findHead:
    #     count=0
    #     for cell in sheet[headingRow]:
    #         st.write(cell.value)
    #         if (cell.value):

    #             if "qty" in cell.value.lower().strip() or "quantity" in cell.value.lower().strip():
    #                 find_qty=True
    #                 qty_index=count
                
    #             if "name" in cell.value.lower().strip() or "product" in cell.value.lower().strip():
    #                 name_index=count
                    
    #             st.write(cell.value)
    #             head.append(cell.value.strip())
    #             count+=1
    nullHead=[]
    oCount=0
    if findHead:
        nullCount=0
        count=0
        for cell in sheet[headingRow]:
            # st.write(cell.value)
            
            
            if (cell.value):
                if "qty" in cell.value.lower().strip() or "quantity" in cell.value.lower().strip():
                    find_qty=True
                    qty_index=oCount
                
                if "name" in cell.value.lower().strip() or "product" in cell.value.lower().strip() or "item" in cell.value.lower().strip():
                    name_index=oCount
                
                
                if "small" in cell.value.lower().strip():
                    small_index=oCount
                if "large" in cell.value.lower().strip():
                    large_index=oCount
                if "medium" in cell.value.lower().strip():
                    medium_index=oCount
                if "xl" in cell.value.lower().strip():
                    xl_index=oCount
                if "uni." in cell.value.lower().strip() or "universal" in cell.value.lower().strip():
                    uni_index=oCount
                    
                    
                oCount+=1
                head.append(cell.value.strip())
            else:
                # head.append(cell.value)
                nullHead.append(count)
                # nullCount+=1
            count+=1
            # st.write(cell.value)
            
            # count+=1
                
        # st.write(head)
        leng=len(head)+len(nullHead)
        # st.write(leng)
        blank=True
        # st.write(nullHead)
        product_rows=[]
        for row in sheet.iter_rows(min_row=headingRow + 1):
            row_data=[]
            for i in range (0,leng):
                if i not in nullHead:
                    # st.write(i)
                    if row[i].value != None :
                        row_data.append(str(row[i].value).strip())
                    else:
                        row_data.append(row[i].value)
            product_rows.append(row_data) 
            
        
        # st.write(product_rows) 
        
        df = pd.DataFrame(product_rows,columns=head)
        rowThank =df[df.apply(lambda row: row.astype(str).str.lower().str.strip().str.contains('thanking').any(), axis=1)].index

        if not rowThank.empty:
            # Get the index of the first occurrence of 'than'
            # st.write(rowThank)
            row_index = rowThank[0]
            # Drop rows below the 'thank' row (including the 'than' row itself)
            df = df.iloc[:row_index]

        

        df = df.dropna(how='all')
        # st.write(df)
        # columns_of_interest = df.columns.tolist()

# Drop rows where all values are NaN or '-'
        # df = df.dropna(how='all', subset=columns_of_interest).loc[~(df[columns_of_interest] == '-').all(axis=1)]    
        
        for index, row in df.iterrows():
            if isinstance(row[head[0]], str) and all(pd.isna(row[col]) for col in head[1:]):
                df = df.drop(index)
        df[df.columns[0]] = range(1, len(df) + 1)
        # st.write("ad:",df[df.columns[name_index]])
        countStr=1
        if "*" in str(df[df.columns[name_index]]):
            findStar=True
            for i in range(len(df)):
                if "*" in str(df.iloc[i, name_index]):
                    # st.write("star found")
                    df.iloc[i, 0] = None
                    continue
                df.iloc[i, 0] = countStr
                countStr+=1
                # st.write(df.iloc[i, name_index])


            
        df = df.reset_index(drop=True)
        df = df.replace('-', None)
        # Display the DataFrame
        # st.write(df)
        edited_df = st.data_editor(df, hide_index=True,width=1500)
        # st.write(uni_index,small_index,medium_index,large_index,xl_index)
        if st.button("Save Changes"):
                # st.write(edited_df)
                # st.write(edited_df.columns[qty_index])
                if not findStar:
                    # edited_df = edited_df[(edited_df[edited_df.columns[qty_index]].notna()) & (edited_df[edited_df.columns[qty_index]] != 0)]
                    # st.write(edited_df.columns[qty_index])
                    edited_df = edited_df.dropna(subset=[edited_df.columns[qty_index]])
                    edited_df = edited_df.dropna(axis=1, how='all')
                    edited_df[edited_df.columns[0]] = range(1, len(edited_df) + 1)
                    # st.write(edited_df)
            
                if findStar:
                    # edited_df = edited_df.dropna(axis=1, how='all').loc[:, (edited_df != 0).any(axis=0)]

                    delete_indices = []
                    stack = []  # temporary stack to store row indices of the current category
                    quantity_found = False
                    # st.write("fi:",edited_df   )
                    # Iterate through DataFrame
                    for index, row in edited_df.iterrows():
                        # st.write(row)
                        # Detect start of a new category by checking 'Serial No.'
                        if pd.notna(row[edited_df.columns[0]]) :
                            # Check the previous category: if no quantity was found, mark all rows in the stack for deletion
                            if stack and not quantity_found:
                                delete_indices.extend(stack)
                            
                            # Reset for the new category
                            stack = [index]
                            quantity_found = False  # Reset quantity_found for new category
                        else:
                            # If it's a continuation of the current category, add index to stack
                            stack.append(index)
                        
                        # If a quantity is found in this row, set quantity_found to True
                        if qty_index:
                            if pd.notna(row[edited_df.columns[qty_index]]) and   row[edited_df.columns[qty_index]] != 0:
                                quantity_found = True
                            # st.write(stack)
                        else:
                            if small_index and medium_index and large_index and xl_index and uni_index:
                                if pd.notna(row[edited_df.columns[small_index]] or row[edited_df.columns[medium_index]] or row[edited_df.columns[large_index] ] or row[edited_df.columns[xl_index]] or row[edited_df.columns[uni_index]]) and  (row[edited_df.columns[small_index]] != 0 or row[edited_df.columns[medium_index]] != 0 or row[edited_df.columns[large_index]] != 0 or row[edited_df.columns[xl_index]] != 0 or row[edited_df.columns[uni_index]] != 0):
                                    quantity_found = True
                            else:
                                if pd.notna( row[edited_df.columns[medium_index]] or row[edited_df.columns[large_index] ]) and  (row[edited_df.columns[medium_index]] != 0 or row[edited_df.columns[large_index]] != 0 or row[edited_df.columns[xl_index]] != 0 ):
                                    quantity_found = True
                    # Final check for the last category
                    if stack and not quantity_found:
                        delete_indices.extend(stack)

                    # Drop rows where entire categories had no quantity
                    edited_df.drop(delete_indices, inplace=True)

                    # Reset index and print the updated DataFrame
                    edited_df.reset_index(drop=True, inplace=True)
                    # st.write("Filtered DataFrame:")
                    # st.write(edited_df)
                    
                    if findStar:
                        countStr=1  
                        for i in range(len(edited_df)):
                            if "*" in str(edited_df.iloc[i, name_index]):
                                # st.write("star found")
                                edited_df.iloc[i, 0] = None
                                continue
                            edited_df.iloc[i, 0] = countStr
                            countStr+=1
                            # st.write(edited_df.iloc[i, name_index])
                    edited_df = edited_df.dropna(axis=1, how='all').loc[:, (edited_df != 0).any(axis=0)]
                
                excel_file = 'header_template.xlsx'  # Replace with your file path
                
                workbook = load_workbook(excel_file)

                # Assuming data is in the first sheet (index 0)
                sheet = workbook.worksheets[0]

                # st.write(edited_df.columns.tolist())
                # Iterate through rows starting from row 15 (index 14 in Python)
                data_to_write = edited_df.values.tolist()
                font_style = Font(name='Arial', size=12, bold=True)
                fill_style = PatternFill(start_color='C4BD97', end_color='C4BD97', fill_type='solid')
                thankFont=Font(name='Arial', size=12, bold=True,italic=True)
                red_font = Font(color="FF0000", size=12, bold=True)
                # align_style = Alignment(horizontal='center', vertical='center')
                align_left = Alignment(horizontal='left', vertical='center')
                # Write data starting from row 15
                for col_idx, header_value in enumerate(edited_df.columns.tolist(), start=1):
                    cell = sheet.cell(row=15, column=col_idx)
                    if "none" in header_value.lower().strip():
                        cell.value=""
                    else:   
                        cell.value = header_value
                    cell.font = font_style
                    cell.fill = fill_style
                    cell.alignment=align_left


                # Write data with formatting starting from row 17
                for row_idx, row_data in enumerate(data_to_write, start=17):
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if isinstance(cell_value, str) and cell_value.startswith('*'):
                        # Remove the star and set the remaining value
                            cell.value = cell_value[1:]
                        # Apply red font color
                            cell.font = red_font
                        else:
                        # Keep the original cell value
                            cell.value = cell_value
                        # cell.value = cell_value
                            cell.font = font_style
                        cell.alignment=align_left
                
                cell = sheet.cell(row=row_idx+2, column=1)
                cell.value = "Thanking You."
                cell.font = thankFont
                cell.alignment=align_left
                # for col in sheet.iter_cols(min_row=15, max_row=15):
                #     for cell in col:
                #         cell.alignment = align_style
                
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get the column name
                    for cell in col[14:]:  # Start from row 15 (index 14)
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2  # Adjusted width formula
                    sheet.column_dimensions[column].width = adjusted_width
                
                sheet['G7']=date
                column_G_width = 15  # Adjust as needed based on your content and date display requirements
                sheet.column_dimensions['G'].width = column_G_width
                # st.write(date)
                
                sheet['A8']=tc
                # column_G_width = 15  # Adjust as needed based on your content and date display requirements
                # sheet.column_dimensions['G'].width = column_G_width
                # st.write(date)
                
                sheet['A9']=tl
                sheet['G8']=orderNO
                sheet['A10']=emails
                # column_G_width = 15  # Adjust as needed based on your content and date display requirements
                # sheet.column_dimensions['G'].width = column_G_width
                # st.write(date)
                
                # save_directory=r"D:\PROJECTS\BS file manager\BS-order-manager\data files\Modified Files"
                if getattr(sys, 'frozen', False):
    # Running in a PyInstaller bundle
                    BASE_DIR = os.path.dirname(sys.executable)
                else:
                    # Running in a normal Python environment
                    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
                    
                save_directory=os.path.join(BASE_DIR, "data files\Modified Files")
                    
                output_directory = os.path.join(save_directory, str(date))
                os.makedirs(output_directory, exist_ok=True)
                # st.write(output_directory)
                output_file = os.path.join(output_directory, f'order_{name}_{date}.{extension}')
                pdf_file = os.path.join(os.getcwd(), f"{save_directory}\{date}\order_{name}_{date}.pdf")
                # Save the workbook
                workbook.save(output_file)
                st.success("File Saved Successfully!")
                convert_excel_to_pdf(output_file, pdf_file)
                if pdf_file:
                    with open(pdf_file, "rb") as f:
                        base64_pdf = base64.b64encode(f.read()).decode('utf-8')
                        pdf_display = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="700" height="500" type="application/pdf"></iframe>'
                    st.markdown(pdf_display, unsafe_allow_html=True)
                

    

def main():
    st.title("ORDER MANAGEMENT")
    
    st.sidebar.title("Company Options")
    # directory = r'D:\PROJECTS\BS file manager\BS-order-manager\data files\Purchase Order'
    if getattr(sys, 'frozen', False):
    # Running in a PyInstaller bundle
        BASE_DIR = os.path.dirname(sys.executable)
    else:
        # Running in a normal Python environment
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    directory=os.path.join(BASE_DIR, "data files\Purchase Order")
    file_list = os.listdir(directory)

                

# FROM DIRECTORY
    
    if file_list:
        orderNO=st.sidebar.text_input("Order Number", 1)
        selectedCompany = st.sidebar.selectbox("Select a company", file_list)

        # Date selection
        selectedDate = st.sidebar.date_input("Order date", date.today())
        textCompany=st.sidebar.text_input("Company Name", selectedCompany.split(".")[0])
        textLocation=st.sidebar.text_input("Company Location", " ")
        emails=st.sidebar.text_input("Email Ids", " ")

        # Display email addresses for the selected company
        if selectedCompany in file_list:
            # st.write(selectedCompany)
            filePath = os.path.join(directory, selectedCompany)
            # st.write(filePath)

            pdf_file_path=fetch_products(filePath, selectedDate,selectedCompany,textCompany,textLocation,orderNO,emails)

if __name__ == "__main__":
    main()

