from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import pandas as pd
# import streamlit as st
import os
import sys
from openpyxl import load_workbook
import base64
from openpyxl.styles import Border, Side
from datetime import datetime
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill,Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.units import mm
from reportlab.pdfbase.pdfmetrics import stringWidth
from pathlib import Path




def test(data,pdffile):
    df = pd.DataFrame(data)

    # Margins and page setup
    left_margin = 20 * mm
    right_margin = 20 * mm
    page_width, _ = A4
    usable_width = page_width - left_margin - right_margin

    # PDF setup
    # pdf_file = "aaaa.pdf"
    doc = SimpleDocTemplate(
        str(pdffile),
        pagesize=A4,
        leftMargin=left_margin,
        rightMargin=right_margin,
    )
    elements = []

    # Font setup
    base_font = "Helvetica"
    bold_font = "Helvetica-Bold"
    italic_font = "Helvetica-Oblique"
    bolditalic_font = "Helvetica-BoldOblique"

    min_font_size = 3
    max_font_size = 10
    column_spacing = 2 * mm  # Only between columns

    num_cols = len(df.columns)
    spacer_cols = num_cols - 1
    total_spacer_width = spacer_cols * column_spacing

    # Step 1: Find width needed per column based on longest content
    col_text_widths = []
    col_max_texts = []

    for col in df.columns:
        values = df[col].dropna().astype(str).tolist() + [col]
        longest_val = max(values, key=lambda x: len(x))
        col_max_texts.append(longest_val)
        col_text_widths.append(len(longest_val))

    # Step 2: Compute proportional column widths (based on text lengths)
    total_chars = sum(col_text_widths)
    available_column_width = usable_width - total_spacer_width
    real_col_widths = [(w / total_chars) * available_column_width for w in col_text_widths]

    # Step 3: Determine maximum font size that fits all max-texts
    uniform_font_size = max_font_size
    while uniform_font_size >= min_font_size:
        fits_all = True
        for i in range(num_cols):
            if stringWidth(col_max_texts[i], base_font, uniform_font_size) > real_col_widths[i]:
                fits_all = False
                break
        if fits_all:
            break
        uniform_font_size -= 0.5

    # Step 4: Define styles
    def make_style(name, fontName=base_font, fontSize=uniform_font_size, color=colors.black):
        return ParagraphStyle(
            name=name,
            fontName=fontName,
            fontSize=fontSize,
            textColor=color,
            leading=fontSize + 1,
            leftIndent=0,
            rightIndent=0,
            wordWrap=None,
            spaceAfter=0,
            spaceBefore=0,
            splitLongWords=False
        )

    styles = {
        'header': make_style("HeaderStyle", fontName=bold_font),
        'data': make_style("DataStyle"),
        'red': make_style("RedStyle", fontName=bold_font, color=colors.red)
    }

    # Step 5: Construct full table data with spacer columns
    def interleave_spacers(row_data):
        result = []
        for i, item in enumerate(row_data):
            result.append(item)
            if i < len(row_data) - 1:
                result.append('')  # spacer cell
        return result

    # Header row
    headers = [Paragraph(col, styles['header']) for col in df.columns]
    table_data = [interleave_spacers(headers)]

    # Data rows
    for row in df.itertuples(index=False):
        row_cells = []
        for i, val in enumerate(row):
            if pd.isna(val) or val is None:
                row_cells.append("")
                continue
            text = str(val).replace('\n', ' ').replace('\r', '')
            is_red = '*' in text
            text = text.replace('*', '')
            style = styles['red'] if is_red else styles['data']

            if stringWidth(text, style.fontName, uniform_font_size) > real_col_widths[i]:
                ellipsis = '...'
                for j in range(len(text), 0, -1):
                    trial = text[:j] + ellipsis
                    if stringWidth(trial, style.fontName, uniform_font_size) <= real_col_widths[i]:
                        text = trial
                        break

            row_cells.append(Paragraph(text, style))
        table_data.append(interleave_spacers(row_cells))

    # Step 6: Interleaved column widths
    final_col_widths = []
    for i in range(num_cols):
        final_col_widths.append(real_col_widths[i])
        if i < num_cols - 1:
            final_col_widths.append(column_spacing)

    # Step 7: Table with styles
    table = Table(table_data, colWidths=final_col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C4BD97')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))

    # Final output
    elements.append(Spacer(1, 187))
    elements.append(table)
    elements.append(Spacer(1, 24))
    elements.append(Paragraph("Thanking You.", make_style("Thank", fontName=bolditalic_font, fontSize=12)))


    
    def draw_header(canvas, doc):
        width, height = A4
        left_margin = 20 * mm
        top_margin = height - 20 * mm

        # Title
        canvas.setFont("Helvetica-BoldOblique", 24)
        canvas.setFillColor(colors.black)
        canvas.drawString(left_margin, top_margin, "Balaji Surgicals")

        # Subheading
        canvas.setFont("Helvetica-Oblique", 16)
        canvas.drawString(left_margin, top_margin - 20, "Nawabi Road, Haldwani–263 139")

        # Orange line
        canvas.setStrokeColor(colors.HexColor("#BF6004"))
        canvas.setLineWidth(4)
        canvas.line(left_margin, top_margin - 24, (width - left_margin)/2+25, top_margin - 24)

        # Left details
        y = top_margin - 40
        row_height = 18
        canvas.setFont("Helvetica", 10)
        canvas.drawString(left_margin, y, "GSTIN No. 05ADGPA2715B1ZH")
        y -= row_height
        canvas.drawString(left_margin, y, "D.L.No. 20B:121663, 21B:121664")

        # Right contact info
        canvas.drawRightString(width - left_margin, top_margin - 40, "Ph.:  97193-04441")
        canvas.drawRightString(width - left_margin, top_margin - 58, "e-mail : balajisumit@yahoo.co.in")

        # Right title
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawRightString(width - left_margin, top_margin, "Order Form")

        # Recipient info
        y -= 2 * row_height
        canvas.setFont("Helvetica-BoldOblique", 12)
        canvas.drawString(left_margin, y, "To,")
        y -= row_height
        canvas.drawString(left_margin, y, "Company Name")
        y -= row_height
        canvas.drawString(left_margin, y, "Company Location")
        y -= row_height
        canvas.setFont("Helvetica", 11)
        canvas.drawString(left_margin, y, "Email ids")

        # Order details
        today = datetime.today().strftime("%d/%m/%y")

        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawRightString(width - left_margin, y + 2 * row_height, f"Date : {today}")
        canvas.drawRightString(width - left_margin, y + row_height, "Order No. : 1")

        # Message
        y -= 2 * row_height
        canvas.setFont("Helvetica-BoldOblique", 12)
        canvas.drawString(left_margin, y, "Dear Sir,")
        y -= row_height
        canvas.drawString(left_margin, y, "Kindly supply the following items:")
        
    doc.build(elements, onFirstPage=draw_header)

# --- ✅ Build Final PDF ---


def fetch_products(file,pdffile):
    # print("hi")
    # os.startfile(file)
    # name,extension=company.split(".")
    workbook = load_workbook(file)
    
    sheet_names = workbook.sheetnames 
    # st.write(sheet_names)
    # st.write(f"Sheet names: {sheet_names}")

    if len(sheet_names) > 1:
        for sheet_name in sheet_names[1:]:
            del workbook[sheet_name]

    sheet_names = workbook.sheetnames 
    # st.write(f"Remaining sheet names: {sheet_names}")

    sheet = workbook[sheet_names[0]]

    findHead=False
    find_qty = False
    qty_index = None
    serial_count = 0
    blank_line = 0
    find_date=False
    date_column=None
    headingRow=None
    name_index=None
    countStr=None
    findStar=False
    
    small_index=None
    medium_index=None
    large_index=None
    xl_index=None
    uni_index=None
    
    head=[]

    # Find 'Qty.' column
    for row in sheet.iter_rows():
        for cell in row:
            # st.write(cell.value)
            if cell.value is not None and isinstance(cell.value, str) and ("qty" in cell.value.strip().lower() or "name" in cell.value.strip().lower() or "size" in cell.value.strip().lower()):
                # st.write(cell.value)
                findHead=True
                
                # find_qty = True
                # qty_column = cell.column
                headingRow=cell.row 
                # st.write(headingRow)
                # st.write(f"'qty' found at Row: {cell.row}, Column: {cell.column}")
                break
        if findHead:
            break

    nullHead=[]
    oCount=0
    if findHead:
        nullCount=0
        count=0
        for cell in sheet[headingRow]:
            # st.write(cell.value)
            
            
            if (cell.value):
                if "qty" in cell.value.lower().strip() or "quantity" in cell.value.lower().strip():
                    find_qty=True
                    qty_index=oCount
                
                if "name" in cell.value.lower().strip() or "product" in cell.value.lower().strip() or "item" in cell.value.lower().strip():
                    name_index=oCount
                
                
                if "small" in cell.value.lower().strip():
                    small_index=oCount
                if "large" in cell.value.lower().strip():
                    large_index=oCount
                if "medium" in cell.value.lower().strip():
                    medium_index=oCount
                if "xl" in cell.value.lower().strip():
                    xl_index=oCount
                if "uni." in cell.value.lower().strip() or "universal" in cell.value.lower().strip():
                    uni_index=oCount
                    
                    
                oCount+=1
                head.append(cell.value.strip())
            else:
                # head.append(cell.value)
                nullHead.append(count)
                # nullCount+=1
            count+=1
            # st.write(cell.value)
            
            # count+=1
                
        # st.write(head)
        leng=len(head)+len(nullHead)
        # st.write(leng)
        blank=True
        # st.write(nullHead)
        product_rows=[]
        for row in sheet.iter_rows(min_row=headingRow + 1):
            row_data=[]
            for i in range (0,leng):
                if i not in nullHead:
                    # st.write(i)
                    if row[i].value != None :
                        row_data.append(str(row[i].value).strip())
                    else:
                        row_data.append(row[i].value)
            product_rows.append(row_data) 
            
        
        # st.write(product_rows) 
        
        df = pd.DataFrame(product_rows,columns=head)
        rowThank =df[df.apply(lambda row: row.astype(str).str.lower().str.strip().str.contains('thanking').any(), axis=1)].index

        if not rowThank.empty:
            # Get the index of the first occurrence of 'than'
            # st.write(rowThank)
            row_index = rowThank[0]
            # Drop rows below the 'thank' row (including the 'than' row itself)
            df = df.iloc[:row_index]

        

        df = df.dropna(how='all')
        # st.write(df)
        # columns_of_interest = df.columns.tolist()

# Drop rows where all values are NaN or '-'
        # df = df.dropna(how='all', subset=columns_of_interest).loc[~(df[columns_of_interest] == '-').all(axis=1)]    
        
        for index, row in df.iterrows():
            if isinstance(row[head[0]], str) and all(pd.isna(row[col]) for col in head[1:]):
                df = df.drop(index)
        df[df.columns[0]] = range(1, len(df) + 1)
        # st.write("ad:",df[df.columns[name_index]])
        countStr=1
        if "*" in str(df[df.columns[name_index]]):
            findStar=True
            for i in range(len(df)):
                if "*" in str(df.iloc[i, name_index]):
                    # st.write("star found")
                    df.iloc[i, 0] = None
                    continue
                df.iloc[i, 0] = countStr
                countStr+=1
                # st.write(df.iloc[i, name_index])


            
        df = df.reset_index(drop=True)
        df = df.replace('-', None)
        # Display the DataFrame
        # st.write(df)
        edited_df = df
        # st.write(uni_index,small_index,medium_index,large_index,xl_index)
        if True:
                # st.write(edited_df)
                # st.write(edited_df.columns[qty_index])
                if not findStar:
                    # edited_df = edited_df[(edited_df[edited_df.columns[qty_index]].notna()) & (edited_df[edited_df.columns[qty_index]] != 0)]
                    # st.write(edited_df.columns[qty_index])
                    edited_df = edited_df.dropna(subset=[edited_df.columns[qty_index]])
                    edited_df = edited_df.dropna(axis=1, how='all')
                    edited_df[edited_df.columns[0]] = range(1, len(edited_df) + 1)
                    # st.write(edited_df)
            
                if findStar:
                    # edited_df = edited_df.dropna(axis=1, how='all').loc[:, (edited_df != 0).any(axis=0)]

                    delete_indices = []
                    stack = []  # temporary stack to store row indices of the current category
                    quantity_found = False
                    # st.write("fi:",edited_df   )
                    # Iterate through DataFrame
                    for index, row in edited_df.iterrows():
                        # st.write(row)
                        # Detect start of a new category by checking 'Serial No.'
                        if pd.notna(row[edited_df.columns[0]]) :
                            # Check the previous category: if no quantity was found, mark all rows in the stack for deletion
                            if stack and not quantity_found:
                                delete_indices.extend(stack)
                            
                            # Reset for the new category
                            stack = [index]
                            quantity_found = False  # Reset quantity_found for new category
                        else:
                            # If it's a continuation of the current category, add index to stack
                            stack.append(index)
                        
                        # If a quantity is found in this row, set quantity_found to True
                        if qty_index:
                            if pd.notna(row[edited_df.columns[qty_index]]) and   row[edited_df.columns[qty_index]] != 0:
                                quantity_found = True
                            # st.write(stack)
                        else:
                            if small_index and medium_index and large_index and xl_index and uni_index:
                                if pd.notna(row[edited_df.columns[small_index]] or row[edited_df.columns[medium_index]] or row[edited_df.columns[large_index] ] or row[edited_df.columns[xl_index]] or row[edited_df.columns[uni_index]]) and  (row[edited_df.columns[small_index]] != 0 or row[edited_df.columns[medium_index]] != 0 or row[edited_df.columns[large_index]] != 0 or row[edited_df.columns[xl_index]] != 0 or row[edited_df.columns[uni_index]] != 0):
                                    quantity_found = True
                            else:
                                if pd.notna( row[edited_df.columns[medium_index]] or row[edited_df.columns[large_index] ]) and  (row[edited_df.columns[medium_index]] != 0 or row[edited_df.columns[large_index]] != 0 or row[edited_df.columns[xl_index]] != 0 ):
                                    quantity_found = True
                    # Final check for the last category
                    if stack and not quantity_found:
                        delete_indices.extend(stack)

                    # Drop rows where entire categories had no quantity
                    edited_df.drop(delete_indices, inplace=True)

                    # Reset index and print the updated DataFrame
                    edited_df.reset_index(drop=True, inplace=True)
                    # st.write("Filtered DataFrame:")
                    # st.write(edited_df)
                    
                    if findStar:
                        countStr=1  
                        for i in range(len(edited_df)):
                            if "*" in str(edited_df.iloc[i, name_index]):
                                # st.write("star found")
                                edited_df.iloc[i, 0] = None
                                continue
                            edited_df.iloc[i, 0] = countStr
                            countStr+=1
                    edited_df = edited_df.dropna(axis=1, how='all').loc[:, (edited_df != 0).any(axis=0)]

                test(edited_df,pdffile)

def main():
   
    
    file = Path(sys.argv[1])
    output_file = Path(sys.argv[2])
    pdf_filename = output_file
    fetch_products(file,pdf_filename)

if __name__ == "__main__":
    main()

