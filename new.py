import sys
import os
import pandas as pd
import base64
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QTableWidget,
    QTableWidgetItem, QLabel, QHBoxLayout, QLineEdit, QMessageBox
)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelEditor(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Order Editor")
        self.setGeometry(100, 100, 1200, 700)
        self.df = pd.DataFrame()

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.init_ui()

    def init_ui(self):
        input_layout = QHBoxLayout()

        self.date_input = QLineEdit()
        self.date_input.setPlaceholderText("Date")

        self.company_input = QLineEdit()
        self.company_input.setPlaceholderText("Company Filename")

        self.tc_input = QLineEdit()
        self.tc_input.setPlaceholderText("TC")

        self.tl_input = QLineEdit()
        self.tl_input.setPlaceholderText("TL")

        self.order_input = QLineEdit()
        self.order_input.setPlaceholderText("Order No.")

        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("Emails")

        input_layout.addWidget(self.date_input)
        input_layout.addWidget(self.company_input)
        input_layout.addWidget(self.tc_input)
        input_layout.addWidget(self.tl_input)
        input_layout.addWidget(self.order_input)
        input_layout.addWidget(self.email_input)

        self.layout.addLayout(input_layout)

        load_button = QPushButton("Load Excel File")
        load_button.clicked.connect(self.load_excel)

        self.layout.addWidget(load_button)

        self.table = QTableWidget()
        self.layout.addWidget(self.table)

        save_button = QPushButton("Save to Excel")
        save_button.clicked.connect(self.save_excel)
        self.layout.addWidget(save_button)

    def load_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls)")
        if not file_path:
            return

        try:
            self.df = self.parse_excel(file_path)
            self.display_df(self.df)
            QMessageBox.information(self, "Success", "Excel file loaded successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def display_df(self, df):
        self.table.setRowCount(df.shape[0])
        self.table.setColumnCount(df.shape[1])
        self.table.setHorizontalHeaderLabels(df.columns.tolist())

        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                value = str(df.iat[i, j]) if pd.notnull(df.iat[i, j]) else ""
                self.table.setItem(i, j, QTableWidgetItem(value))

    def get_table_data(self):
        rows, cols = self.table.rowCount(), self.table.columnCount()
        data = []
        headers = [self.table.horizontalHeaderItem(i).text() for i in range(cols)]
        for i in range(rows):
            row_data = []
            for j in range(cols):
                item = self.table.item(i, j)
                row_data.append(item.text() if item else None)
            data.append(row_data)
        return pd.DataFrame(data, columns=headers)

    def save_excel(self):
        edited_df = self.get_table_data()
        try:
            file_path = QFileDialog.getOpenFileName(self, "Select Template Excel File", "", "Excel Files (*.xlsx *.xls)")[0]
            if not file_path:
                return

            workbook = load_workbook(file_path)
            sheet = workbook.worksheets[0]

            # Clear previous data from row 15 onwards
            for row in sheet.iter_rows(min_row=15, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None

            font_style = Font(name='Arial', size=12, bold=True)
            fill_style = PatternFill(start_color='C4BD97', end_color='C4BD97', fill_type='solid')
            align_left = Alignment(horizontal='left', vertical='center')

            for col_idx, header_value in enumerate(edited_df.columns.tolist(), start=1):
                cell = sheet.cell(row=15, column=col_idx)
                cell.value = header_value
                cell.font = font_style
                cell.fill = fill_style
                cell.alignment = align_left

            for row_idx, row_data in enumerate(edited_df.values.tolist(), start=17):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = cell_value
                    cell.font = font_style
                    cell.alignment = align_left

            # Additional info from UI
            sheet['G7'] = self.date_input.text()
            sheet['A8'] = self.tc_input.text()
            sheet['A9'] = self.tl_input.text()
            sheet['G8'] = self.order_input.text()
            sheet['A10'] = self.email_input.text()

            save_path, _ = QFileDialog.getSaveFileName(self, "Save Final Excel", "", "Excel Files (*.xlsx)")
            if not save_path:
                return

            workbook.save(save_path)
            QMessageBox.information(self, "Success", f"Saved to {save_path}")

        except Exception as e:
            QMessageBox.critical(self, "Save Error", str(e))

    def parse_excel(self, file):
        wb = load_workbook(file)
        sheet = wb.active

        headingRow = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and (
                        "qty" in cell.value.lower() or "name" in cell.value.lower()):
                    headingRow = cell.row
                    break
            if headingRow:
                break

        if headingRow is None:
            raise ValueError("Heading row not found.")

        headers = [cell.value if cell.value else f"Column{i}" for i, cell in enumerate(sheet[headingRow])]
        data = []

        for row in sheet.iter_rows(min_row=headingRow + 1):
            row_data = []
            for i, cell in enumerate(row[:len(headers)]):
                row_data.append(str(cell.value).strip() if cell.value is not None else "")
            data.append(row_data)

        df = pd.DataFrame(data, columns=headers)
        df = df.dropna(how='all')
        return df

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelEditor()
    window.show()
    sys.exit(app.exec_())