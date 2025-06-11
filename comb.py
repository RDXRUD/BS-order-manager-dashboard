import sys
import os
import pandas as pd
from datetime import date
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QVBoxLayout, QHBoxLayout, QLineEdit,
    QPushButton, QComboBox, QDateEdit, QFileDialog, QTableWidget,
    QTableWidgetItem, QSplitter, QMessageBox, QCompleter
)
from PyQt5.QtCore import Qt, QDate
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelOrderEditor(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Combined Excel Order Management")
        self.setGeometry(100, 100, 1400, 800)
        self.df = pd.DataFrame()
        self.setup_ui()

    def setup_ui(self):
        main_layout = QHBoxLayout()
        splitter = QSplitter(Qt.Horizontal)

        # === Left Form Panel ===
        form_widget = QWidget()
        form_layout = QVBoxLayout()

        # Order Number
        self.order_no_input = QLineEdit("1")
        form_layout.addLayout(self._make_form_row("Order Number:", self.order_no_input))

        # File Directory and File List
        base_dir = os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__))
        self.directory = os.path.join(base_dir, "data files", "Purchase Order")
        self.file_list = [f for f in os.listdir(self.directory) if f.endswith('.xlsx') or f.endswith('.xls')]

        # Company File Dropdown
        self.company_combo = QComboBox()
        self.company_combo.setEditable(True)
        self.company_combo.addItems(self.file_list)
        completer = QCompleter(self.file_list)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.company_combo.setCompleter(completer)
        self.company_combo.lineEdit().editingFinished.connect(
            lambda: self.handle_manual_entry(self.company_combo.currentText())
        )
        self.company_combo.currentTextChanged.connect(self.update_company_name)
        self.company_combo.currentTextChanged.connect(self.load_excel)
        form_layout.addLayout(self._make_form_row("Select a Company:", self.company_combo))

        # Order Date
        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        form_layout.addLayout(self._make_form_row("Order Date:", self.date_edit))

        # Company Name
        self.company_name_input = QLineEdit()
        form_layout.addLayout(self._make_form_row("Company Name:", self.company_name_input))

        # Location
        self.location_input = QLineEdit()
        form_layout.addLayout(self._make_form_row("Company Location:", self.location_input))

        # Emails
        self.email_input = QLineEdit()
        form_layout.addLayout(self._make_form_row("Email IDs:", self.email_input))

        # Save Button
        save_button = QPushButton("Save to Excel")
        save_button.clicked.connect(self.save_excel)
        form_layout.addWidget(save_button)
        form_layout.addStretch()

        form_widget.setLayout(form_layout)
        splitter.addWidget(form_widget)

        # === Right Table Panel ===
        self.table = QTableWidget()
        splitter.addWidget(self.table)
        splitter.setSizes([400, 1000])

        main_layout.addWidget(splitter)
        self.setLayout(main_layout)

        # Auto load first file if exists
        if self.file_list:
            self.update_company_name(self.file_list[0])
            self.load_excel(self.file_list[0])

    def _make_form_row(self, label_text, widget):
        layout = QHBoxLayout()
        label = QLabel(label_text)
        label.setMinimumWidth(140)
        layout.addWidget(label)
        layout.addWidget(widget)
        return layout

    def handle_manual_entry(self, text):
        if text in self.file_list:
            self.company_combo.setCurrentText(text)
            self.update_company_name(text)
            self.load_excel(text)

    def update_company_name(self, selected_file):
        self.company_name_input.setText(selected_file.split(".")[0])

    def load_excel(self, selected_file):
        file_path = os.path.join(self.directory, selected_file)
        try:
            self.df = self.parse_excel(file_path)
            self.display_df(self.df)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def parse_excel(self, file):
        wb = load_workbook(file)
        sheet = wb.active

        headingRow = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and (
                        "qty" in cell.value.lower() or "name" in cell.value.lower()):
                    headingRow = cell.row
                    break
            if headingRow:
                break

        if headingRow is None:
            raise ValueError("Heading row not found.")

        headers = [cell.value if cell.value else f"Column{i}" for i, cell in enumerate(sheet[headingRow])]
        data = []

        for row in sheet.iter_rows(min_row=headingRow + 1):
            row_data = [str(cell.value).strip() if cell.value is not None else "" for cell in row[:len(headers)]]
            data.append(row_data)

        df = pd.DataFrame(data, columns=headers)
        df = df.dropna(how='all')
        return df

    def display_df(self, df):
        self.table.setRowCount(df.shape[0])
        self.table.setColumnCount(df.shape[1])
        self.table.setHorizontalHeaderLabels(df.columns.tolist())

        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                value = str(df.iat[i, j]) if pd.notnull(df.iat[i, j]) else ""
                self.table.setItem(i, j, QTableWidgetItem(value))

    def get_table_data(self):
        rows, cols = self.table.rowCount(), self.table.columnCount()
        data = []
        headers = [self.table.horizontalHeaderItem(i).text() for i in range(cols)]
        for i in range(rows):
            row_data = []
            for j in range(cols):
                item = self.table.item(i, j)
                row_data.append(item.text() if item else None)
            data.append(row_data)
        return pd.DataFrame(data, columns=headers)

    def save_excel(self):
        edited_df = self.get_table_data()
        try:
            file_path = os.path.join(self.directory, self.company_combo.currentText())
            workbook = load_workbook(file_path)
            sheet = workbook.worksheets[0]

            for row in sheet.iter_rows(min_row=15, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None

            font_style = Font(name='Arial', size=12, bold=True)
            fill_style = PatternFill(start_color='C4BD97', end_color='C4BD97', fill_type='solid')
            align_left = Alignment(horizontal='left', vertical='center')

            for col_idx, header_value in enumerate(edited_df.columns.tolist(), start=1):
                cell = sheet.cell(row=15, column=col_idx)
                cell.value = header_value
                cell.font = font_style
                cell.fill = fill_style
                cell.alignment = align_left

            for row_idx, row_data in enumerate(edited_df.values.tolist(), start=17):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = cell_value
                    cell.font = font_style
                    cell.alignment = align_left

            sheet['G7'] = self.date_edit.text()
            sheet['A8'] = self.company_name_input.text()
            sheet['A9'] = self.location_input.text()
            sheet['G8'] = self.order_no_input.text()
            sheet['A10'] = self.email_input.text()

            save_path, _ = QFileDialog.getSaveFileName(self, "Save Final Excel", "", "Excel Files (*.xlsx)")
            if not save_path:
                return

            workbook.save(save_path)
            QMessageBox.information(self, "Success", f"Saved to {save_path}")

        except Exception as e:
            QMessageBox.critical(self, "Save Error", str(e))

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelOrderEditor()
    window.show()
    sys.exit(app.exec_())