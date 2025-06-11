import streamlit as st
import streamlit.web.cli as stcli
import os
import sys
from openpyxl import load_workbook
import win32com.client
import pythoncom
import base64
from openpyxl.styles import Border, Side
from datetime import date
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill,Alignment
from openpyxl.utils import get_column_letter
import streamlit.runtime.scriptrunner.magic_funcs


def resolve_path(path):
    resolved_path = os.path.abspath(os.path.join(os.getcwd(), path))
    return resolved_path


if __name__ == "__main__":
    sys.argv = [
        "streamlit",
        "run",
        resolve_path("st.py"),
        "--global.developmentMode=false",
    ]
    sys.exit(stcli.main())